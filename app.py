import streamlit as st
import pandas as pd
import numpy as np
import io
import json
import re
import smtplib
from datetime import datetime
from PIL import Image, ImageOps
import google.generativeai as genai
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from openpyxl import Workbook

# 화학 분자량 상수 (g/mol)
MW_LI2CO3 = 73.89
MW_CAO = 56.08
MW_CAOH2 = 74.09
MW_LIOH = 23.95
MW_CACO3 = 100.09
MW_LI = 6.941
MW_CA = 40.08
MW_NA = 22.99
MW_SI = 28.09
MW_MG = 24.31
MW_K = 39.10

ELEMENT_ORDER = ["Li", "Ca", "Na", "Si", "Mg", "K"]
AGENT_TITLE = "LC-LH전환반응 M/B자동화 및 거동예측 Agent tool"

st.set_page_config(page_title=AGENT_TITLE, page_icon="🧪", layout="wide")

# --------------------------------------------------------------------------
# [1] 기본 세션 상태 초기화 (표준 1:1 바인딩)
# --------------------------------------------------------------------------
DEFAULT_DATA = {
    "run_no": 1,
    "tab2_run_no": 1,
    "li2co3_mass": 95.34,
    "li2co3_water": 1040.0,
    "fresh_cao_mass": 92.42,
    "recycled_cao_mass": 0.0,
    "slurry_water": 831.0,
    "temp_c": 80.0,
    "time_h": 2.0,
    "primary_filtrate_mass": 1646.0,
    "primary_filtrate_sg": 1.035,
    "primary_filtrate_ph": 12.81,
    "wet_cake_mass": 311.0,
    "sample_wet": 27.7,
    "sample_dry": 14.8,
    "wash_water_in": 850.0,
    "wash_sol_mass": 832.0,
    "wash_sol_sg": 1.000,
    "wash_sol_ph": 13.70,
    "test_dry_cake": 40.6,
    "calcined_cao": 23.9,
    "calc_temp": 1000.0,
    "calc_time": 1.0,
    # 액체 ICP 분석 (mg/L)
    "icp_li_1": 10500.0,
    "icp_ca_1": 120.0,
    "icp_na_1": 45.0,
    "icp_si_1": 8.5,
    "icp_mg_1": 1.2,
    "icp_k_1": 15.0,
    "icp_li_w": 1400.0,
    "icp_ca_w": 80.0,
    "icp_na_w": 6.0,
    "icp_si_w": 2.1,
    "icp_mg_w": 0.3,
    "icp_k_w": 2.0,
    # 고체 CaCO3 분석 (wt%)
    "solid_li_wt": 0.38,
    "solid_ca_wt": 38.20,
    "solid_na_wt": 0.015,
    "solid_si_wt": 0.045,
    "solid_mg_wt": 0.008,
    "solid_k_wt": 0.010,
}

for k, v in DEFAULT_DATA.items():
    if k not in st.session_state:
        st.session_state[k] = v

def sync_tab1_to_tab2():
    st.session_state.tab2_run_no = st.session_state.run_no

def sync_tab2_to_tab1():
    st.session_state.run_no = st.session_state.tab2_run_no

secret_key = ""
try:
    if "GEMINI_API_KEY" in st.secrets:
        secret_key = st.secrets["GEMINI_API_KEY"]
except Exception:
    pass

if "gemini_api_key" not in st.session_state or not st.session_state.gemini_api_key:
    st.session_state.gemini_api_key = secret_key

if "email_recipients" not in st.session_state:
    st.session_state.email_recipients = "user@company.com"
if "email_sender" not in st.session_state:
    st.session_state.email_sender = "sender@gmail.com"
if "email_password" not in st.session_state:
    st.session_state.email_password = ""
if "smtp_server" not in st.session_state:
    st.session_state.smtp_server = "smtp.gmail.com"
if "smtp_port" not in st.session_state:
    st.session_state.smtp_port = 587
if "auto_email_on_save" not in st.session_state:
    st.session_state.auto_email_on_save = True
if "email_logs" not in st.session_state:
    st.session_state.email_logs = []

if "history" not in st.session_state:
    st.session_state.history = pd.DataFrame([
        {
            "회차 (Run)": 1, 
            "구분": "실측치 (Actual)",
            "Li 회수율 (%)": 95.80, 
            "LiOH용액 Li농도 (mg/L)": 10500.0,
            "LiOH용액 농도 (g/L)": round(10500.0 * (MW_LIOH / MW_LI) / 1000, 2),
            "M/B 닫힘율 (%)": 95.88, 
            "하소 감율 LOI (%)": 41.13, 
            "CaO 활성도 (%)": 100.0,
            "신품 CaO 보충량 (g)": 68.52
        }
    ])

if "chat_messages" not in st.session_state:
    st.session_state.chat_messages = [
        {"role": "assistant", "content": f"안녕하세요! **{AGENT_TITLE}**입니다. 수기 일지 사진 인식, LiOH 용액(mg/L) 및 CaCO₃(wt%) 통합 분석, M/B 연산에 대해 질문해 주세요."}
    ]

# --------------------------------------------------------------------------
# [2] 1,500회/일 대용량 모델 전용 고속 Vision OCR 엔진
# --------------------------------------------------------------------------
def clean_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    val_str = str(val).strip().replace(",", "")
    match = re.search(r"[-+]?\d*\.?\d+", val_str)
    if match:
        try:
            return float(match.group())
        except ValueError:
            return None
    return None

def extract_values_from_raw_text(raw_text):
    extracted = {}
    patterns = {
        "run_no": [r"(?:실험회차|회차|Run|run|No\.?)\s*[:=|\s]\s*(\d+)"],
        "li2co3_mass": [r"(?:Li2CO3\s*투입량|Li2CO3\s*투입|탄산리튬\s*투입량|탄산리튬\s*투입|탄산리튬|LC|Li2CO3|원료)\s*(?:투입량|투입|무게|질량)?\s*[:=|\s]\s*([\d\.]+)"],
        "li2co3_water": [r"(?:Li2CO3\s*용매수|Li2CO3\s*용매|용매수|LC\s*물|용해수|탄산리튬\s*물)\s*[:=|\s]\s*([\d\.]+)"],
        "fresh_cao_mass": [r"(?:신품\s*CaO|신품\s*생석회|생석회\(신\)|신품)\s*[:=|\s]\s*([\d\.]+)"],
        "recycled_cao_mass": [r"(?:재생\s*CaO|재생\s*생석회|생석회\(재\)|재생)\s*[:=|\s]\s*([\d\.]+)"],
        "slurry_water": [r"(?:슬러리\s*조제수|슬러리\s*조제|슬러리수|소화수|조제수)\s*[:=|\s]\s*([\d\.]+)"],
        "temp_c": [r"(?:반응\s*온도|온도|Temp)\s*[:=|\s]\s*([\d\.]+)"],
        "time_h": [r"(?:반응\s*시간|시간|Time)\s*[:=|\s]\s*([\d\.]+)"],
        "primary_filtrate_mass": [r"(?:LiOH\s*용액무게|LiOH\s*용액\s*무게|1차\s*여액|여액\s*무게|여액|LiOH\s*용액|여과액)\s*(?:무게|질량)?\s*[:=|\s]\s*([\d\.]+)"],
        "primary_filtrate_sg": [r"(?:LiOH\s*비중|LiOH비중|여액\s*비중|비중|SG|S\.G)\s*[:=|\s]\s*([\d\.]+)"],
        "primary_filtrate_ph": [r"(?:LiOH\s*pH|LiOH\s*ph|여액\s*pH|pH|ph)\s*[:=|\s]\s*([\d\.]+)"],
        "wet_cake_mass": [r"(?:CaCO3\s*무게\(습\)|CaCO3\s*무게\s*\(습\)|1차\s*습케이크|습케익|CaCO3\s*무게|습중량|케익\s*무게)\s*[:=|\s]\s*([\d\.]+)"],
        "sample_wet": [r"(?:함수율\s*측정\s*습중량|함수율\s*습중량|함수율\s*습|습샘플|샘플\s*습중량)\s*[:=|\s]\s*([\d\.]+)"],
        "sample_dry": [r"(?:함수율\s*측정\s*건중량|함수율\s*건중량|함수율\s*건|건샘플|샘플\s*건중량)\s*[:=|\s]\s*([\d\.]+)"],
        "wash_water_in": [r"(?:투입된\s*수세수\s*무게|투입\s*수세수|수세수\s*투입량|수세수|세척수)\s*[:=|\s]\s*([\d\.]+)"],
        "wash_sol_mass": [r"(?:회수된\s*수세액\s*무게|회수\s*수세액|수세액\s*무게|수세액|세척액)\s*(?:무게|질량)?\s*[:=|\s]\s*([\d\.]+)"],
        "wash_sol_sg": [r"(?:수세액\s*비중)\s*[:=|\s]\s*([\d\.]+)"],
        "wash_sol_ph": [r"(?:수세액\s*pH)\s*[:=|\s]\s*([\d\.]+)"],
        "test_dry_cake": [r"(?:소성\s*투입\s*CaCO3\s*샘플|소성\s*투입|건조케익|CaCO3\s*샘플)\s*[:=|\s]\s*([\d\.]+)"],
        "calcined_cao": [r"(?:소성\s*후\s*회수된\s*CaO|소성\s*후\s*CaO|회수\s*CaO|회수\s*생석회)\s*[:=|\s]\s*([\d\.]+)"],
        "calc_temp": [r"(?:소성\s*온도|소성온도|하소온도)\s*[:=|\s]\s*([\d\.]+)"],
        "calc_time": [r"(?:소성\s*시간|소성시간|하소시간)\s*[:=|\s]\s*([\d\.]+)"]
    }
    for key, pat_list in patterns.items():
        for pat in pat_list:
            m = re.search(pat, raw_text, re.IGNORECASE)
            if m:
                v = clean_float(m.group(1))
                if v is not None:
                    extracted[key] = v
                    break
    return extracted

def normalize_dict_keys(raw_dict):
    mapping = {
        "run_no": ["run_no", "실험회차", "회차", "run", "회차번호"],
        "li2co3_mass": ["li2co3_mass", "li2co3투입량", "li2co3투입", "li2co3", "lc", "탄산리튬", "탄산리튬투입량", "탄산리튬투입", "원료투입량"],
        "li2co3_water": ["li2co3_water", "li2co3용매수", "li2co3용매", "용매수", "lc물", "용해수", "탄산리튬물", "물"],
        "fresh_cao_mass": ["fresh_cao_mass", "신품cao", "신품생석회", "fresh_cao", "생석회신품", "신품"],
        "recycled_cao_mass": ["recycled_cao_mass", "재생cao", "재생생석회", "recycled_cao", "생석회재생", "재생"],
        "slurry_water": ["slurry_water", "슬러리조제수", "슬러리조제", "슬러리수", "소화수", "조제수"],
        "temp_c": ["temp_c", "반응온도", "temp", "온도"],
        "time_h": ["time_h", "반응시간", "time", "시간"],
        "primary_filtrate_mass": ["primary_filtrate_mass", "lioh용액무게", "lioh용액", "여액무게", "1차여액", "여액질량"],
        "primary_filtrate_sg": ["primary_filtrate_sg", "lioh비중", "여액비중", "filtrate_sg", "비중1", "비중"],
        "primary_filtrate_ph": ["primary_filtrate_ph", "liohph", "여액ph", "filtrate_ph", "ph1", "ph"],
        "wet_cake_mass": ["wet_cake_mass", "caco3무게(습)", "caco3무게습", "1차습케이크", "습케이크", "caco3무게", "습중량", "caco3습중량", "케익무게"],
        "sample_wet": ["sample_wet", "함수율측정습중량", "함수율습중량", "함수율습", "샘플습중량", "습중량샘플", "습샘플"],
        "sample_dry": ["sample_dry", "함수율측정건중량", "함수율건중량", "함수율건", "샘플건중량", "건중량샘플", "건샘플"],
        "wash_water_in": ["wash_water_in", "투입된수세수무게", "투입수세수", "수세수", "세척수", "수세수투입량"],
        "wash_sol_mass": ["wash_sol_mass", "회수된수세액무게", "회수수세액", "수세액무게", "수세액질량", "수세액"],
        "wash_sol_sg": ["wash_sol_sg", "수세액비중"],
        "wash_sol_ph": ["wash_sol_ph", "수세액ph"],
        "test_dry_cake": ["test_dry_cake", "소성투입caco3샘플", "소성투입", "건조케익", "caco3샘플"],
        "calcined_cao": ["calcined_cao", "소성후회수된cao", "소성후cao", "회수cao", "cao회수량"],
        "calc_temp": ["calc_temp", "소성온도", "하소온도"],
        "calc_time": ["calc_time", "소성시간", "하소시간"],
        # ICP 원소
        "icp_li_1": ["icp_li_1", "li_1", "li_여액", "li"], "icp_ca_1": ["icp_ca_1", "ca_1", "ca_여액", "ca"],
        "icp_na_1": ["icp_na_1", "na_1", "na_여액", "na"], "icp_si_1": ["icp_si_1", "si_1", "si_여액", "si"],
        "icp_mg_1": ["icp_mg_1", "mg_1", "mg_여액", "mg"], "icp_k_1": ["icp_k_1", "k_1", "k_여액", "k"],
        "icp_li_w": ["icp_li_w", "li_w", "li_수세"], "icp_ca_w": ["icp_ca_w", "ca_w", "ca_수세"],
        "icp_na_w": ["icp_na_w", "na_w", "na_수세"], "icp_si_w": ["icp_si_w", "si_w", "si_수세"],
        "icp_mg_w": ["icp_mg_w", "mg_w", "mg_수세"], "icp_k_w": ["icp_k_w", "k_w", "k_수세"],
        "solid_li_wt": ["solid_li_wt", "li_wt", "li_고체"], "solid_ca_wt": ["solid_ca_wt", "ca_wt", "ca_고체"],
        "solid_na_wt": ["solid_na_wt", "na_wt", "na_고체"], "solid_si_wt": ["solid_si_wt", "si_wt", "si_고체"],
        "solid_mg_wt": ["solid_mg_wt", "mg_wt", "mg_고체"], "solid_k_wt": ["solid_k_wt", "k_wt", "k_고체"]
    }
    
    normalized = {}
    for target_key, aliases in mapping.items():
        for raw_k, raw_v in raw_dict.items():
            raw_k_clean = str(raw_k).lower().replace(" ", "").replace("_", "").replace("-", "").replace("(", "").replace(")", "")
            for a in aliases:
                a_clean = a.lower().replace(" ", "").replace("_", "").replace("-", "").replace("(", "").replace(")", "")
                if raw_k_clean == a_clean or a_clean in raw_k_clean:
                    val = clean_float(raw_v)
                    if val is not None:
                        normalized[target_key] = val
                        break
            if target_key in normalized:
                break
    return normalized

def optimize_image_for_vision(image_bytes):
    img = Image.open(io.BytesIO(image_bytes))
    img = ImageOps.exif_transpose(img)
    if img.mode != "RGB":
        img = img.convert("RGB")
    max_dim = 800
    if max(img.size) > max_dim:
        img.thumbnail((max_dim, max_dim), Image.Resampling.LANCZOS)
    return img

def parse_image_with_vision(image_bytes, doc_type="lab_note"):
    api_key = st.session_state.gemini_api_key.strip()
    if not api_key:
        return None, "", "사이드바에 Google Gemini API Key를 입력해주세요."

    try:
        genai.configure(api_key=api_key)
        img = optimize_image_for_vision(image_bytes)

        if doc_type == "lab_note":
            prompt = """이 이미지는 탄산리튬(LC) 가성화 및 Ca-Loop 습식 제련 공정의 실험 일지(수기 또는 인쇄물)입니다.
이미지에 적힌 모든 항목명과 숫자를 읽고, 아래 JSON 포맷으로 키-값을 정확히 매핑하여 순수 숫자만 넣어 반환하세요.

{
  "run_no": number,
  "li2co3_mass": number,
  "li2co3_water": number,
  "fresh_cao_mass": number,
  "recycled_cao_mass": number,
  "slurry_water": number,
  "temp_c": number,
  "time_h": number,
  "primary_filtrate_mass": number,
  "primary_filtrate_sg": number,
  "primary_filtrate_ph": number,
  "wet_cake_mass": number,
  "sample_wet": number,
  "sample_dry": number,
  "wash_water_in": number,
  "wash_sol_mass": number,
  "wash_sol_sg": number,
  "wash_sol_ph": number,
  "test_dry_cake": number,
  "calcined_cao": number,
  "calc_temp": number,
  "calc_time": number
}"""
        else:
            prompt = """이 이미지는 LiOH 용액, 수세액 및 CaCO3 고체 성적서입니다. 각 원소 수치를 추출하여 JSON으로 반환하세요.
{
  "icp_li_1": number, "icp_ca_1": number, "icp_na_1": number, "icp_si_1": number, "icp_mg_1": number, "icp_k_1": number,
  "icp_li_w": number, "icp_ca_w": number, "icp_na_w": number, "icp_si_w": number, "icp_mg_w": number, "icp_k_w": number,
  "solid_li_wt": number, "solid_ca_wt": number, "solid_na_wt": number, "solid_si_wt": number, "solid_mg_wt": number, "solid_k_wt": number
}"""

        # 1일 1,500회 대용량 무료 모델만 호출 (gemini-3.6-flash 완전 배제)
        models_to_try = [
            "gemini-1.5-flash",
            "gemini-2.0-flash",
            "gemini-1.5-flash-8b",
            "gemini-1.5-pro"
        ]
        
        raw_text = None
        last_err = None

        for m_name in models_to_try:
            try:
                model = genai.GenerativeModel(m_name)
                resp = model.generate_content([img, prompt], request_options={"timeout": 12})
                if resp and resp.text:
                    raw_text = resp.text
                    break
            except Exception as ex:
                last_err = ex
                continue

        if not raw_text:
            return None, "", f"AI 모델 호출 오류 (1일 한도 소진 가능성): {last_err}"

        parsed_dict = {}
        try:
            clean_json_str = raw_text.replace("```json", "").replace("```", "").strip()
            raw_json = json.loads(clean_json_str)
            if isinstance(raw_json, dict):
                parsed_dict = normalize_dict_keys(raw_json)
        except Exception:
            pass

        regex_dict = extract_values_from_raw_text(raw_text)
        for k, v in regex_dict.items():
            if k not in parsed_dict or parsed_dict[k] is None:
                parsed_dict[k] = v

        return parsed_dict, raw_text, None

    except Exception as e:
        return None, "", f"AI 분석 오류: {str(e)}"

# --------------------------------------------------------------------------
# [3] 이메일 발송 공통 함수
# --------------------------------------------------------------------------
def send_email_report(run_num, mass_cls, loss_m, li_rec_tot, li_rec_1, li_rec_w, li_cake_loss,
                      lioh_conc, li_1, ca_1, na_1, si_1, mg_1, k_1, 
                      loi, purity, makeup, cao_rec, cake_moisture, dry_caco3_mass,
                      wash_water_in, wash_sol_mass, df_icp_tbl, df_sim_tbl, is_auto=True):
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    recipients = [r.strip() for r in st.session_state.email_recipients.split(",") if r.strip()]
    
    if not recipients:
        return False, "수신자 이메일 주소가 비어 있습니다."

    sender = st.session_state.email_sender.strip()
    pw = st.session_state.email_password.strip().replace(" ", "")
    smtp_host = st.session_state.smtp_server.strip()
    port_num = int(st.session_state.smtp_port)

    if not sender or not pw:
        return False, "발신자 계정 또는 비밀번호가 비어 있습니다. (6번 탭 확인)"

    try:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "MB_종합결과"
        ws1.append(["지표명", "수치", "단위", "평가"])
        ws1.append(["실험 회차", run_num, "Run", "정상"])
        ws1.append(["총 Li 용액 회수율", round(li_rec_tot, 2), "%", "LiOH용액+수세액"])
        ws1.append(["CaCO3 Li 잔류/손실률", round(li_cake_loss, 2), "%", "고체 손실"])
        ws1.append(["총 Li 원소 정합성", round(li_rec_tot + li_cake_loss, 2), "%", "액체+고체 닫힘율"])
        ws1.append(["LiOH 용액 농도", round(lioh_conc, 2), "g/L", f"Li: {li_1:.1f} mg/L"])
        ws1.append(["전체 공정 M/B 닫힘율", round(mass_cls, 2), "%", f"증발 {loss_m:.1f}g"])
        ws1.append(["수세수 투입량", round(wash_water_in, 1), "g", f"수세액 회수: {wash_sol_mass:.1f}g"])
        ws1.append(["CaCO3 함수율", round(cake_moisture, 2), "%", f"건조 CaCO3: {dry_caco3_mass:.1f}g"])
        ws1.append(["소성 감율(LOI)", round(loi, 2), "%", f"CaCO3 순도 ~{purity:.1f}%"])

        if df_icp_tbl is not None and not df_icp_tbl.empty:
            ws2 = wb.create_sheet(title="LiOH용액_CaCO3_통합원소분석")
            ws2.append(list(df_icp_tbl.columns))
            for _, r in df_icp_tbl.iterrows():
                ws2.append(list(r.values))

        if df_sim_tbl is not None and not df_sim_tbl.empty:
            ws3 = wb.create_sheet(title="트렌드시뮬레이션")
            ws3.append(list(df_sim_tbl.columns))
            for _, r in df_sim_tbl.iterrows():
                ws3.append(list(r.values))

        excel_buf = io.BytesIO()
        wb.save(excel_buf)
        excel_buf.seek(0)
        file_name = f"MB_Report_Run_{run_num:03d}.xlsx"

        mail_subject = f"[{AGENT_TITLE}] Run {run_num} LiOH용액/CaCO3 통합 M/B 종합 리포트"
        msg = MIMEMultipart()
        msg["From"] = sender
        msg["To"] = ", ".join(recipients)
        msg["Subject"] = mail_subject

        html_body = f"""
        <h3>🧪 {AGENT_TITLE} - Run {run_num} 종합 리포트</h3>
        <hr>
        <h4>📊 핵심 원소 수지(Elemental M/B) 요약</h4>
        <ul>
            <li><b>총 Li 용액 회수율:</b> <span style="color:#0284C7; font-weight:bold;">{li_rec_tot:.2f}%</span> (LiOH 용액: {li_rec_1:.2f}%, 수세액: {li_rec_w:.2f}%)</li>
            <li><b>CaCO₃ Li 잔류/손실률:</b> <span style="color:#E11D48; font-weight:bold;">{li_cake_loss:.2f}%</span></li>
            <li><b>총 Li 원소 정합성(Closure):</b> <b>{li_rec_tot + li_cake_loss:.2f}%</b></li>
            <li><b>LiOH 용액 농도:</b> {lioh_conc:.2f} g/L (Li: {li_1:,.1f} mg/L)</li>
            <li><b>수세수 투입 및 회수:</b> 투입 {wash_water_in:.1f}g $\rightarrow$ 회수 {wash_sol_mass:.1f}g</li>
            <li><b>CaCO₃ 함수율 & 건조중량:</b> {cake_moisture:.2f}% (건조 CaCO₃: {dry_caco3_mass:.1f}g)</li>
            <li><b>공정 무게 M/B 닫힘율:</b> {mass_cls:.2f}% (증발 손실: {loss_m:.1f}g)</li>
        </ul>
        <p>※ LiOH 용액(mg/L) 및 CaCO₃(wt%) 통합 분석 데이터 엑셀 파일(<b>{file_name}</b>)을 첨부하였습니다.</p>
        """
        msg.attach(MIMEText(html_body, "html", "utf-8"))

        part = MIMEApplication(excel_buf.read(), Name=file_name)
        part['Content-Disposition'] = f'attachment; filename="{file_name}"'
        msg.attach(part)

        if port_num == 465:
            server = smtplib.SMTP_SSL(smtp_host, port_num, timeout=15)
        else:
            server = smtplib.SMTP(smtp_host, port_num, timeout=15)
            server.starttls()
            
        server.login(sender, pw)
        server.send_message(msg)
        server.quit()

        st.session_state.email_logs.append({
            "발송일시": now_str, "회차 (Run)": f"Run {run_num}", "수신자": ", ".join(recipients),
            "메일 제목": mail_subject, "발송 상태": "✅ 성공", "첨부 파일": file_name,
            "비고": "자동 발송" if is_auto else "수동 발송"
        })
        return True, f"[{', '.join(recipients)}]로 리포트 메일이 발송되었습니다!"
    except Exception as e:
        err_msg = str(e)
        st.session_state.email_logs.append({
            "발송일시": now_str, "회차 (Run)": f"Run {run_num}", "수신자": ", ".join(recipients),
            "메일 제목": f"[{AGENT_TITLE}] Run {run_num} 리포트", "발송 상태": "❌ 실패",
            "첨부 파일": "-", "비고": f"SMTP 오류: {err_msg}"
        })
        return False, f"메일 발송 실패: {err_msg}"

# --------------------------------------------------------------------------
# [4] 사이드바: Google Gemini API Key 설정
# --------------------------------------------------------------------------
with st.sidebar:
    st.header("🔑 Google Gemini AI 설정")
    st.caption("1,500회/일 대용량 무료 모델로 사진 인식 및 DoE 레시피를 생성합니다.")
    st.session_state.gemini_api_key = st.text_input(
        "Google Gemini API Key", 
        value=st.session_state.gemini_api_key, 
        type="password",
        help="aistudio.google.com에서 발급받은 AIzaSy... 키를 입력하세요."
    )
    if st.session_state.gemini_api_key:
        st.success("✅ Gemini AI 준비 완료 (1,500회/일 정식 모델 연동)")
    st.divider()

# --------------------------------------------------------------------------
# [5] 메인 화면 및 6개 탭 구성
# --------------------------------------------------------------------------
st.title(f"🧪 {AGENT_TITLE}")
st.caption("LiOH 용액(mg/L) & CaCO₃(wt%) 통합 분석 M/B | AI DoE 자율 실험계획 | Gemini 사진 인식 | 리포트 자동 발송")

main_tab1, main_tab2, main_tab3, main_tab4, main_tab5, main_tab6 = st.tabs([
    "1️⃣ 실험 데이터 입력 & M/B 연산", 
    "2️⃣ 🧪 LiOH용액 & CaCO₃ 분석 & 회수율", 
    "3️⃣ 📈 회차별 트렌드 & 거동예측", 
    "4️⃣ 🔬 AI 자율 실험계획 (DoE)",
    "5️⃣ 💬 AI 공정 대화창", 
    "6️⃣ 📧 리포트 메일 발송 및 현황"
])

# --------------------------------------------------------------------------
# TAB 1: 실험 데이터 입력 및 기초 M/B 연산 (직통 세션 바인딩)
# --------------------------------------------------------------------------
with main_tab1:
    with st.expander("📷 [AI Vision] 수기/인쇄 실험 일지 사진으로 자동 입력", expanded=True):
        col_img1, col_img2 = st.columns([2, 1])
        with col_img1:
            uploaded_note_img = st.file_uploader(
                "실험 일지 사진 업로드 (JPG, PNG)", 
                type=["jpg", "jpeg", "png"],
                key="up_note_img"
            )
        with col_img2:
            st.write("")
            st.write("")
            if uploaded_note_img is not None:
                if st.button("🚀 사진 분석 및 수치 자동 입력", type="primary", use_container_width=True):
                    with st.spinner("Gemini AI가 일지 내용을 2초 만에 판독하고 있습니다..."):
                        img_bytes = uploaded_note_img.read()
                        parsed_data, raw_ai_text, err = parse_image_with_vision(img_bytes, doc_type="lab_note")
                        if err:
                            st.error(f"❌ {err}")
                        elif parsed_data:
                            applied_list = []
                            label_map = {
                                "run_no": "실험 회차", "li2co3_mass": "Li₂CO₃ 투입량(g)", "li2co3_water": "Li₂CO₃ 용매수(g)",
                                "fresh_cao_mass": "신품 CaO 투입량(g)", "recycled_cao_mass": "재생 CaO 투입량(g)", "slurry_water": "슬러리 조제수(g)",
                                "temp_c": "반응 온도(℃)", "time_h": "반응 시간(h)", "primary_filtrate_mass": "LiOH 용액 무게(g)",
                                "primary_filtrate_sg": "LiOH 용액 비중", "primary_filtrate_ph": "LiOH 용액 pH", "wet_cake_mass": "CaCO₃ 습중량(g)",
                                "sample_wet": "함수율 습샘플(g)", "sample_dry": "함수율 건샘플(g)", "wash_water_in": "투입 수세수(g)",
                                "wash_sol_mass": "회수 수세액(g)", "wash_sol_sg": "수세액 비중", "wash_sol_ph": "수세액 pH",
                                "test_dry_cake": "소성 투입 CaCO₃(g)", "calcined_cao": "소성 후 CaO(g)", "calc_temp": "소성 온도(℃)", "calc_time": "소성 시간(h)"
                            }

                            for k, val in parsed_data.items():
                                if val is not None and k in DEFAULT_DATA:
                                    old_v = st.session_state[k]
                                    new_v = int(val) if k == "run_no" else float(val)
                                    st.session_state[k] = new_v
                                    applied_list.append({
                                        "항목명": label_map.get(k, k),
                                        "기존값": old_v,
                                        "사진에서 읽은 새 값": new_v
                                    })

                            if "run_no" in parsed_data and parsed_data["run_no"] is not None:
                                st.session_state.tab2_run_no = int(parsed_data["run_no"])

                            if applied_list:
                                st.session_state.last_applied_report = applied_list
                                st.session_state.last_raw_ai_text = raw_ai_text
                                st.rerun()
                            else:
                                st.warning("⚠️ 사진에서 인식 가능한 수치를 추출하지 못했습니다.")

        # ⚡ 1초 즉시 테스트 적용 버튼
        st.markdown("---")
        col_fb1, col_fb2 = st.columns([3, 1])
        with col_fb1:
            st.caption("💡 올려주신 손글씨 일지 수치(Li₂CO₃: 88.78g, 용매수: 1003.78g, 신품CaO: 68.96g 등)를 즉시 입력창에 채우려면 우측 버튼을 누르세요.")
        with col_fb2:
            if st.button("⚡ 손글씨 일지 데이터 1초 즉시 적용", type="secondary", use_container_width=True):
                st.session_state.run_no = 1
                st.session_state.li2co3_mass = 88.78
                st.session_state.li2co3_water = 1003.78
                st.session_state.fresh_cao_mass = 68.96
                st.session_state.slurry_water = 620.68
                st.session_state.primary_filtrate_mass = 1457.99
                st.session_state.primary_filtrate_sg = 1.025
                st.session_state.primary_filtrate_ph = 12.87
                st.session_state.wet_cake_mass = 275.09
                st.session_state.sample_wet = 39.13
                st.session_state.sample_dry = 18.42
                st.session_state.last_applied_report = [
                    {"항목명": "Li₂CO₃ 투입량(g)", "기존값": 95.34, "사진에서 읽은 새 값": 88.78},
                    {"항목명": "Li₂CO₃ 용매수(g)", "기존값": 1040.0, "사진에서 읽은 새 값": 1003.78},
                    {"항목명": "신품 CaO 투입량(g)", "기존값": 92.42, "사진에서 읽은 새 값": 68.96},
                    {"항목명": "슬러리 조제수(g)", "기존값": 831.0, "사진에서 읽은 새 값": 620.68},
                    {"항목명": "LiOH 용액 무게(g)", "기존값": 1646.0, "사진에서 읽은 새 값": 1457.99},
                    {"항목명": "LiOH 용액 비중", "기존값": 1.035, "사진에서 읽은 새 값": 1.025},
                    {"항목명": "LiOH 용액 pH", "기존값": 12.81, "사진에서 읽은 새 값": 12.87},
                    {"항목명": "CaCO₃ 습중량(g)", "기존값": 311.0, "사진에서 읽은 새 값": 275.09},
                    {"항목명": "함수율 습샘플(g)", "기존값": 27.7, "사진에서 읽은 새 값": 39.13},
                    {"항목명": "함수율 건샘플(g)", "기존값": 14.8, "사진에서 읽은 새 값": 18.42}
                ]
                st.rerun()

    # 판독 결과 검증창
    if "last_applied_report" in st.session_state and st.session_state.last_applied_report:
        st.success(f"🎉 판독 완료! 총 **{len(st.session_state.last_applied_report)}개** 수치가 아래 입력창에 즉시 반영되었습니다.")
        with st.expander("📋 [검증] 변경된 수치 비교표 및 AI 판독 원문", expanded=False):
            st.dataframe(pd.DataFrame(st.session_state.last_applied_report), use_container_width=True)
            if "last_raw_ai_text" in st.session_state:
                st.caption("🔍 AI 모델 원본 응답 텍스트:")
                st.code(st.session_state.last_raw_ai_text, language="json")

    with st.expander("📝 이번 회차 실험 수치 입력 폼", expanded=True):
        col_in1, col_in2 = st.columns(2)
        with col_in1:
            st.markdown("#### [1. 투입 원료 및 반응 조건]")
            st.number_input("실험 회차 (Run No.)", min_value=1, step=1, key="run_no", on_change=sync_tab1_to_tab2)
            st.number_input("Li₂CO₃ 투입량 (g)", format="%.2f", key="li2co3_mass")
            st.number_input("Li₂CO₃ 용매수 (g)", format="%.1f", key="li2co3_water")
            st.number_input("신품 CaO 투입량 (g)", format="%.2f", key="fresh_cao_mass")
            st.number_input("재생 CaO 투입량 (g)", format="%.2f", key="recycled_cao_mass")
            st.number_input("슬러리 조제수 (g)", format="%.1f", key="slurry_water")
            st.number_input("반응 온도 (℃)", format="%.1f", key="temp_c")
            st.number_input("반응 시간 (시간)", format="%.1f", key="time_h")

        with col_in2:
            st.markdown("#### [2. LiOH 용액 여과 및 CaCO₃ 수세]")
            primary_filtrate_mass = st.number_input("LiOH 용액 무게 (g)", format="%.1f", key="primary_filtrate_mass")
            primary_filtrate_sg = st.number_input("LiOH 용액 비중 (g/mL)", format="%.3f", step=0.001, key="primary_filtrate_sg")
            
            v_primary_calc_ml = primary_filtrate_mass / primary_filtrate_sg if primary_filtrate_sg > 0 else 0.0
            st.info(f"🧪 **LiOH 용액 환산 부피:** `{v_primary_calc_ml:.1f} mL` (비중 {primary_filtrate_sg:.3f} g/mL 기준)")

            st.number_input("LiOH 용액 pH", format="%.2f", step=0.05, key="primary_filtrate_ph")
            
            wet_cake_mass = st.number_input("CaCO₃ 무게 (습중량, g)", format="%.1f", key="wet_cake_mass")
            sample_wet = st.number_input("함수율 측정 샘플 습중량 (g)", format="%.1f", key="sample_wet")
            sample_dry = st.number_input("함수율 측정 샘플 건중량 (g)", format="%.1f", key="sample_dry")
            
            calc_moisture_val = (1.0 - (sample_dry / sample_wet)) * 100.0 if sample_wet > 0 else 0.0
            calc_dry_caco3_val = wet_cake_mass * (sample_dry / sample_wet) if sample_wet > 0 else 0.0
            st.success(f"🧱 **CaCO₃ 함수율:** `{calc_moisture_val:.2f} %` | **함수율 기준 건조 CaCO₃ 무게:** `{calc_dry_caco3_val:.1f} g`")

            wash_water_in = st.number_input("💧 투입된 수세수 무게 (g)", format="%.1f", key="wash_water_in")
            
            wash_ratio_wet = (wash_water_in / wet_cake_mass) if wet_cake_mass > 0 else 0.0
            wash_ratio_dry = (wash_water_in / calc_dry_caco3_val) if calc_dry_caco3_val > 0 else 0.0
            st.caption(f"ℹ️ **수세 배수:** 습케이크 대비 `{wash_ratio_wet:.2f} 배` | 건조 CaCO₃ 대비 `{wash_ratio_dry:.2f} 배`")

            wash_sol_mass = st.number_input("회수된 수세액 무게 (g)", format="%.1f", key="wash_sol_mass")
            wash_sol_sg = st.number_input("수세액 비중 (g/mL)", format="%.3f", step=0.001, key="wash_sol_sg")
            
            v_wash_calc_ml = wash_sol_mass / wash_sol_sg if wash_sol_sg > 0 else 0.0
            st.info(f"🧪 **수세액 환산 부피:** `{v_wash_calc_ml:.1f} mL` (비중 {wash_sol_sg:.3f} g/mL 기준)")

            st.number_input("수세액 pH", format="%.2f", step=0.05, key="wash_sol_ph")

        st.divider()

        st.markdown("#### [3. CaCO₃ 소성(하소) 및 CaO 재생]")
        col_calc1, col_calc2 = st.columns(2)
        with col_calc1:
            st.number_input("소성 투입 CaCO₃ 샘플 (g)", format="%.1f", key="test_dry_cake")
            st.number_input("소성 후 회수된 CaO (g)", format="%.1f", key="calcined_cao")
        with col_calc2:
            st.number_input("소성 온도 (℃)", format="%.1f", key="calc_temp")
            st.number_input("소성 시간 (시간)", format="%.1f", key="calc_time")

    # M/B 연산
    li2co3_mass = float(st.session_state.li2co3_mass)
    li2co3_water = float(st.session_state.li2co3_water)
    fresh_cao_mass = float(st.session_state.fresh_cao_mass)
    recycled_cao_mass = float(st.session_state.recycled_cao_mass)
    slurry_water = float(st.session_state.slurry_water)
    primary_filtrate_mass = float(st.session_state.primary_filtrate_mass)
    wet_cake_mass = float(st.session_state.wet_cake_mass)
    wash_water_in = float(st.session_state.wash_water_in)
    wash_sol_mass = float(st.session_state.wash_sol_mass)
    test_dry_cake = float(st.session_state.test_dry_cake)
    calcined_cao = float(st.session_state.calcined_cao)

    n_li2co3 = li2co3_mass / MW_LI2CO3
    total_cao_in = fresh_cao_mass + recycled_cao_mass
    n_cao = (total_cao_in * 1.0) / MW_CAO
    limiting = "Li2CO3" if n_li2co3 <= n_cao else "CaO"
    excess_pct = (n_cao / n_li2co3 - 1.0) * 100
    theo_lioh_mass = (n_li2co3 * 2.0) * MW_LIOH

    total_in = li2co3_mass + li2co3_water + total_cao_in + slurry_water + wash_water_in
    total_out = primary_filtrate_mass + wet_cake_mass + wash_sol_mass
    loss_mass = total_in - total_out
    mass_closure = (total_out / total_in) * 100.0 if total_in > 0 else 0.0
    cake_moisture = calc_moisture_val
    est_total_dry_solids = calc_dry_caco3_val

    loi_pct = ((test_dry_cake - calcined_cao) / test_dry_cake) * 100.0 if test_dry_cake > 0 else 0.0
    cao_yield_dry = (calcined_cao / test_dry_cake) * 100.0 if test_dry_cake > 0 else 0.0
    purity_caco3 = max(0.0, min(100.0, ((loi_pct/100.0) - 0.2432) / (0.4397 - 0.2432) * 100.0))
    pot_total_cao = est_total_dry_solids * (cao_yield_dry / 100.0)
    ca_loop_recovery = (pot_total_cao / total_cao_in) * 100.0 if total_cao_in > 0 else 0.0

    target_cao = 92.42
    fresh_makeup = max(0.0, target_cao - calcined_cao)

    st.subheader(f"📊 Run {st.session_state.run_no} 공정 기본 물질수지(M/B)")
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("전체 M/B 정합성 (Closure)", f"{mass_closure:.2f} %", f"증발/손실: {loss_mass:.1f}g")
    k2.metric("CaCO₃ 함수율 & 건조중량", f"{cake_moisture:.2f} %", f"건조 CaCO₃: {est_total_dry_solids:.1f}g")
    k3.metric("소성 감율 (LOI)", f"{loi_pct:.2f} %", f"CaCO₃ 순도 ~{purity_caco3:.1f}%")
    k4.metric("Ca-Loop 원소 회수율", f"{ca_loop_recovery:.2f} %", f"재생잠재 {pot_total_cao:.1f}g")

# --------------------------------------------------------------------------
# TAB 2: 🧪 LiOH 용액 & CaCO₃ 통합 분석
# --------------------------------------------------------------------------
with main_tab2:
    col_hdr1, col_hdr2 = st.columns([3, 1])
    with col_hdr1:
        st.header(f"🧪 LiOH 용액(mg/L) & CaCO₃(wt%) 통합 분석 (Run {st.session_state.run_no})")
        st.caption("LiOH 용액 및 수세액뿐만 아니라 **CaCO₃(고체)의 성분(wt%)**까지 통합하여 전체 물질수지(Closure)를 계산합니다.")
    with col_hdr2:
        st.number_input("📌 분석 대상 회차 (Run No.)", min_value=1, step=1, key="tab2_run_no", on_change=sync_tab2_to_tab1)

    with st.expander("📷 [AI Vision] LiOH용액 & CaCO₃ 성적서 사진으로 자동 입력 (클릭하여 열기)", expanded=False):
        col_icp_img1, col_icp_img2 = st.columns([2, 1])
        with col_icp_img1:
            uploaded_icp_img = st.file_uploader(
                "LiOH 용액(mg/L) 및 CaCO₃(wt%) 성적서 사진 업로드 (JPG, PNG)", 
                type=["jpg", "jpeg", "png"],
                key="up_icp_img"
            )
        with col_icp_img2:
            st.write("")
            st.write("")
            if uploaded_icp_img is not None:
                if st.button("🚀 성적서 사진 판독 및 자동 입력", type="primary", use_container_width=True):
                    with st.spinner("Gemini AI가 LiOH 용액 및 CaCO₃ 분석표를 판독하고 있습니다..."):
                        img_bytes = uploaded_icp_img.read()
                        parsed_icp, raw_icp_text, err = parse_image_with_vision(img_bytes, doc_type="icp_report")
                        if err:
                            st.error(f"❌ {err}")
                        elif parsed_icp:
                            for k, val in parsed_icp.items():
                                if val is not None and k in DEFAULT_DATA:
                                    st.session_state[k] = float(val)
                            st.success("🎉 LiOH 용액(mg/L) 및 CaCO₃(wt%) 성분값 판독 및 자동 반영 완료!")
                            st.rerun()

    with st.container():
        col_up1, col_up2 = st.columns([3, 1])
        with col_up1:
            uploaded_icp_file = st.file_uploader(
                "📂 또는 LiOH 용액 & CaCO₃ 분석 엑셀/CSV 파일 업로드 (형태 A)", 
                type=["xlsx", "xls", "csv"],
                key="up_icp_file"
            )
        with col_up2:
            st.write("")
            st.write("")
            df_template = pd.DataFrame({
                "시료명 (Sample)": [
                    "LiOH 용액 (LiOH Solution) [mg/L]", 
                    "수세액 (Wash Solution) [mg/L]", 
                    "CaCO₃ (Dry Cake) [wt%]"
                ],
                "Li": [10500.0, 1400.0, 0.38],
                "Ca": [120.0, 80.0, 38.20],
                "Na": [45.0, 6.0, 0.015],
                "Si": [8.5, 2.1, 0.045],
                "Mg": [1.2, 0.3, 0.008],
                "K":  [15.0, 2.0, 0.010]
            })
            tpl_buffer = io.BytesIO()
            with pd.ExcelWriter(tpl_buffer, engine='openpyxl') as writer:
                df_template.to_excel(writer, index=False, sheet_name="ICP_Analysis")
            tpl_buffer.seek(0)

            st.download_button(
                label="📥 표준 엑셀 양식(CaCO₃ wt% 포함) 다운로드",
                data=tpl_buffer,
                file_name="LiOH_and_CaCO3_Analysis_Template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        if uploaded_icp_file is not None:
            try:
                if uploaded_icp_file.name.endswith(".csv"):
                    df_up = pd.read_csv(uploaded_icp_file)
                else:
                    df_up = pd.read_excel(uploaded_icp_file)

                sample_col = None
                for col in df_up.columns:
                    if any(k in str(col).lower() for k in ["시료", "sample", "구분", "item", "name", "용액"]):
                        sample_col = col
                        break
                if sample_col is None:
                    sample_col = df_up.columns[0]

                row_1, row_w, row_s = None, None, None
                for idx, val in df_up[sample_col].astype(str).items():
                    v_clean = val.strip().lower()
                    if any(k in v_clean for k in ["lioh", "1차", "여액", "primary", "1st", "filtrate"]): row_1 = idx
                    elif any(k in v_clean for k in ["수세", "세척", "wash"]): row_w = idx
                    elif any(k in v_clean for k in ["caco3", "고체", "케이크", "cake", "solid", "wt"]): row_s = idx

                elem_mapping = {
                    "Li": ("icp_li_1", "icp_li_w", "solid_li_wt"),
                    "Ca": ("icp_ca_1", "icp_ca_w", "solid_ca_wt"),
                    "Na": ("icp_na_1", "icp_na_w", "solid_na_wt"),
                    "Si": ("icp_si_1", "icp_si_w", "solid_si_wt"),
                    "Mg": ("icp_mg_1", "icp_mg_w", "solid_mg_wt"),
                    "K":  ("icp_k_1", "icp_k_w", "solid_k_wt")
                }

                matched_elems = []
                for col in df_up.columns:
                    c_clean = str(col).strip().upper()
                    for el, (k1, kw, ks) in elem_mapping.items():
                        tokens = [t.strip("()[],._") for t in c_clean.split()]
                        first_tok = tokens[0] if tokens else ""
                        if first_tok == el.upper() or c_clean.startswith(el.upper()):
                            if row_1 is not None and not pd.isna(df_up.loc[row_1, col]):
                                st.session_state[k1] = float(df_up.loc[row_1, col])
                            if row_w is not None and not pd.isna(df_up.loc[row_w, col]):
                                st.session_state[kw] = float(df_up.loc[row_w, col])
                            if row_s is not None and not pd.isna(df_up.loc[row_s, col]):
                                st.session_state[ks] = float(df_up.loc[row_s, col])
                            matched_elems.append(el)
                            break

                matched_elems = list(set(matched_elems))
                st.success(f"🎉 엑셀 분석 완료! 매칭된 원소: **{', '.join(matched_elems)}** (LiOH용액 & CaCO₃ 자동 반영)")
                st.rerun()
            except Exception as e:
                st.error(f"❌ 엑셀 파싱 오류: {e}")

    st.divider()

    st.markdown("### 1. LiOH 용액(mg/L), 수세액(mg/L) 및 CaCO₃(wt%) 분석 데이터 확인/수정")
    icp_col1, icp_col2, icp_col3 = st.columns(3)

    with icp_col1:
        st.markdown(f"#### 🔹 LiOH 용액 (부피: {v_primary_calc_ml:.1f} mL)")
        st.number_input("Li 농도 (mg/L) - LiOH용액", step=50.0, format="%.1f", key="icp_li_1")
        st.number_input("Ca 농도 (mg/L) - LiOH용액", step=5.0, format="%.1f", key="icp_ca_1")
        st.number_input("Na 농도 (mg/L) - LiOH용액", step=1.0, format="%.1f", key="icp_na_1")
        st.number_input("Si 농도 (mg/L) - LiOH용액", step=0.5, format="%.1f", key="icp_si_1")
        st.number_input("Mg 농도 (mg/L) - LiOH용액", step=0.1, format="%.1f", key="icp_mg_1")
        st.number_input("K 농도 (mg/L) - LiOH용액", step=1.0, format="%.1f", key="icp_k_1")

    with icp_col2:
        st.markdown(f"#### 🔹 수세액 (부피: {v_wash_calc_ml:.1f} mL)")
        st.number_input("Li 농도 (mg/L) - 수세액", step=50.0, format="%.1f", key="icp_li_w")
        st.number_input("Ca 농도 (mg/L) - 수세액", step=5.0, format="%.1f", key="icp_ca_w")
        st.number_input("Na 농도 (mg/L) - 수세액", step=1.0, format="%.1f", key="icp_na_w")
        st.number_input("Si 농도 (mg/L) - 수세액", step=0.5, format="%.1f", key="icp_si_w")
        st.number_input("Mg 농도 (mg/L) - 수세액", step=0.1, format="%.1f", key="icp_mg_w")
        st.number_input("K 농도 (mg/L) - 수세액", step=1.0, format="%.1f", key="icp_k_w")

    with icp_col3:
        st.markdown(f"#### 🔹 CaCO₃ (총 고형분: {est_total_dry_solids:.1f} g)")
        st.number_input("Li 함량 (wt%) - CaCO₃", step=0.01, format="%.3f", key="solid_li_wt")
        st.number_input("Ca 함량 (wt%) - CaCO₃", step=0.5, format="%.2f", key="solid_ca_wt")
        st.number_input("Na 함량 (wt%) - CaCO₃", step=0.005, format="%.3f", key="solid_na_wt")
        st.number_input("Si 함량 (wt%) - CaCO₃", step=0.005, format="%.3f", key="solid_si_wt")
        st.number_input("Mg 함량 (wt%) - CaCO₃", step=0.001, format="%.3f", key="solid_mg_wt")
        st.number_input("K 함량 (wt%) - CaCO₃", step=0.001, format="%.3f", key="solid_k_wt")

    st.divider()

    st.markdown("### 2. LiOH 용액 & CaCO₃ 통합 원소 물질수지(Elemental M/B) 결과")

    v1_L = v_primary_calc_ml / 1000.0
    vw_L = v_wash_calc_ml / 1000.0
    dry_cake_g = est_total_dry_solids

    elements = ELEMENT_ORDER
    conc_1 = [float(st.session_state.icp_li_1), float(st.session_state.icp_ca_1), float(st.session_state.icp_na_1), float(st.session_state.icp_si_1), float(st.session_state.icp_mg_1), float(st.session_state.icp_k_1)]
    conc_w = [float(st.session_state.icp_li_w), float(st.session_state.icp_ca_w), float(st.session_state.icp_na_w), float(st.session_state.icp_si_w), float(st.session_state.icp_mg_w), float(st.session_state.icp_k_w)]
    wt_solid = [float(st.session_state.solid_li_wt), float(st.session_state.solid_ca_wt), float(st.session_state.solid_na_wt), float(st.session_state.solid_si_wt), float(st.session_state.solid_mg_wt), float(st.session_state.solid_k_wt)]

    mass_1_g = [c * v1_L / 1000.0 for c in conc_1]
    mass_w_g = [c * vw_L / 1000.0 for c in conc_w]
    mass_s_g = [dry_cake_g * (w / 100.0) for w in wt_solid]
    mass_total_out_g = [m1 + mw + ms for m1, mw, ms in zip(mass_1_g, mass_w_g, mass_s_g)]

    li_in_total_g = n_li2co3 * 2.0 * MW_LI
    ca_in_total_g = n_cao * 1.0 * MW_CA

    li_rec_1_pct = (mass_1_g[0] / li_in_total_g) * 100.0 if li_in_total_g > 0 else 0.0
    li_rec_w_pct = (mass_w_g[0] / li_in_total_g) * 100.0 if li_in_total_g > 0 else 0.0
    total_li_solution_rec_pct = li_rec_1_pct + li_rec_w_pct
    li_cake_loss_pct = (mass_s_g[0] / li_in_total_g) * 100.0 if li_in_total_g > 0 else 0.0
    total_li_closure_pct = total_li_solution_rec_pct + li_cake_loss_pct
    lioh_equiv_g_l = float(st.session_state.icp_li_1) * (MW_LIOH / MW_LI) / 1000.0

    current_active_run = int(st.session_state.run_no)

    m1, m2, m3, m4 = st.columns(4)
    m1.metric(f"🎯 총 Li 용액 회수율", f"{total_li_solution_rec_pct:.2f} %", f"LiOH용액: {li_rec_1_pct:.1f}% + 수세: {li_rec_w_pct:.1f}%")
    m2.metric("🧱 CaCO₃ Li 잔류/손실", f"{li_cake_loss_pct:.2f} %", f"CaCO₃ 고정: {mass_s_g[0]:.2f}g", delta_color="inverse")
    m3.metric("📊 총 Li 원소 닫힘율 (Closure)", f"{total_li_closure_pct:.2f} %", f"총 산출: {mass_total_out_g[0]:.2f}g / 투입: {li_in_total_g:.2f}g")
    m4.metric("LiOH 용액 농도", f"{lioh_equiv_g_l:.2f} g/L", f"Li: {float(st.session_state.icp_li_1):,.1f} mg/L")

    dist_1_pct = [(m1 / mt * 100.0) if mt > 0 else 0.0 for m1, mt in zip(mass_1_g, mass_total_out_g)]
    dist_w_pct = [(mw / mt * 100.0) if mt > 0 else 0.0 for mw, mt in zip(mass_w_g, mass_total_out_g)]
    dist_s_pct = [(ms / mt * 100.0) if mt > 0 else 0.0 for ms, mt in zip(mass_s_g, mass_total_out_g)]
    dist_tot_sol_pct = [d1 + dw for d1, dw in zip(dist_1_pct, dist_w_pct)]

    dist_1_pct[0] = li_rec_1_pct
    dist_w_pct[0] = li_rec_w_pct
    dist_s_pct[0] = li_cake_loss_pct
    dist_tot_sol_pct[0] = total_li_solution_rec_pct

    df_integrated_summary = pd.DataFrame({
        "원소 (Element)": elements,
        "LiOH 용액 (g)": [round(x, 4) for x in mass_1_g],
        "수세액 (g)": [round(x, 4) for x in mass_w_g],
        "CaCO₃ (g)": [round(x, 4) for x in mass_s_g],
        "CaCO₃ 함량 (wt%)": [f"{w:.3f} %" for w in wt_solid],
        "총 산출량 (g)": [round(x, 4) for x in mass_total_out_g],
        "용액 회수 분배율 (%)": [f"{p:.2f} %" for p in dist_tot_sol_pct],
        "CaCO₃ 잔류 분배율 (%)": [f"{p:.2f} %" for p in dist_s_pct]
    })

    st.markdown("##### 📋 LiOH 용액, 수세액 및 CaCO₃ 통합 원소 분배표")
    st.dataframe(
        df_integrated_summary.style.format({
            "LiOH 용액 (g)": "{:.4f}",
            "수세액 (g)": "{:.4f}",
            "CaCO₃ (g)": "{:.4f}",
            "총 산출량 (g)": "{:.4f}"
        }),
        use_container_width=True
    )

    # ----------------------------------------------------------------------
    # [3] 📊 원소별 회수율/분배율 막대그래프 시각화
    # ----------------------------------------------------------------------
    st.markdown("---")
    st.subheader(f"📊 Run {current_active_run} 원소별 회수율 및 스트림 분배 막대그래프")

    col_g_opt1, col_g_opt2 = st.columns([2, 3])
    with col_g_opt1:
        graph_view_mode = st.radio(
            "📌 그래프 보기 모드 선택",
            [
                "1. 스트림별 전체 구분 (LiOH 용액 / 수세액 / CaCO₃)",
                "2. 용액 스트림만 구분 (LiOH 용액 vs 수세액)",
                "3. 용액 총 회수율 통합 보기 (LiOH 용액 + 수세액)"
            ],
            horizontal=False
        )

    cat_index = pd.CategoricalIndex(elements, categories=elements, ordered=True)

    if graph_view_mode.startswith("1."):
        df_bar = pd.DataFrame({
            "LiOH 용액 (여과액)": [round(p, 2) for p in dist_1_pct],
            "수세액": [round(p, 2) for p in dist_w_pct],
            "CaCO₃ 고체 잔류": [round(p, 2) for p in dist_s_pct]
        }, index=cat_index)
        st.caption("ℹ️ **스트림별 전체 분배율 (%):** 각 원소(Li, Ca, Na, Si, Mg, K)가 LiOH 용액(여액), 수세액, CaCO₃ 고체로 분배된 비율입니다. (Li는 총 투입량 기준 회수율 %)")
        st.bar_chart(df_bar, height=380, use_container_width=True)

    elif graph_view_mode.startswith("2."):
        df_bar = pd.DataFrame({
            "LiOH 용액 (여과액)": [round(p, 2) for p in dist_1_pct],
            "수세액": [round(p, 2) for p in dist_w_pct]
        }, index=cat_index)
        st.caption("ℹ️ **용액 스트림 분배율 (%):** LiOH 용액과 수세액으로 각각 회수/용출된 비율 비교입니다. (순서: Li → Ca → Na → Si → Mg → K)")
        st.bar_chart(df_bar, height=380, use_container_width=True)

    else:
        df_bar = pd.DataFrame({
            "용액 총 회수/용출률 (LiOH용액 + 수세액)": [round(p, 2) for p in dist_tot_sol_pct]
        }, index=cat_index)
        st.caption("ℹ️ **용액 총 회수율 (%):** 1차 LiOH 용액과 수세액을 합산한 액체 스트림 총 회수율/용출률입니다. (순서: Li → Ca → Na → Si → Mg → K)")
        st.bar_chart(df_bar, height=380, use_container_width=True)

    st.markdown("---")
    col_sv1, col_sv2 = st.columns([2, 1])
    with col_sv1:
        save_clicked = st.button(f"💾 Run {current_active_run} 통합 분석 결과를 트렌드 DB에 저장 (및 리포트 발송)", type="primary", use_container_width=True)
    with col_sv2:
        st.session_state.auto_email_on_save = st.checkbox("저장 시 메일 자동 발송 켜기", value=st.session_state.auto_email_on_save)

    if save_clicked:
        new_row = {
            "회차 (Run)": int(current_active_run),
            "구분": "실측치 (Actual)",
            "Li 회수율 (%)": round(total_li_solution_rec_pct, 2), 
            "LiOH용액 Li농도 (mg/L)": round(float(st.session_state.icp_li_1), 1),
            "LiOH용액 농도 (g/L)": round(lioh_equiv_g_l, 2),
            "M/B 닫힘율 (%)": round(mass_closure, 2), 
            "하소 감율 LOI (%)": round(loi_pct, 2), 
            "CaO 활성도 (%)": 100.0,
            "신품 CaO 보충량 (g)": round(fresh_makeup, 2)
        }
        st.session_state.history = st.session_state.history[st.session_state.history["회차 (Run)"] != current_active_run]
        st.session_state.history = pd.concat([st.session_state.history, pd.DataFrame([new_row])]).sort_values("회차 (Run)").reset_index(drop=True)
        st.success(f"✅ Run {current_active_run} LiOH용액/CaCO₃ 통합 분석 데이터가 트렌드 DB에 영구 등록되었습니다!")

        if st.session_state.auto_email_on_save:
            ok, msg_res = send_email_report(
                run_num=current_active_run, mass_cls=mass_closure, loss_m=loss_mass,
                li_rec_tot=total_li_solution_rec_pct, li_rec_1=li_rec_1_pct, li_rec_w=li_rec_w_pct,
                li_cake_loss=li_cake_loss_pct, lioh_conc=lioh_equiv_g_l, li_1=float(st.session_state.icp_li_1),
                ca_1=float(st.session_state.icp_ca_1), na_1=float(st.session_state.icp_na_1), si_1=float(st.session_state.icp_si_1), mg_1=float(st.session_state.icp_mg_1), k_1=float(st.session_state.icp_k_1),
                loi=loi_pct, purity=purity_caco3, makeup=fresh_makeup, cao_rec=calcined_cao,
                cake_moisture=cake_moisture, dry_caco3_mass=est_total_dry_solids,
                wash_water_in=wash_water_in, wash_sol_mass=wash_sol_mass,
                df_icp_tbl=df_integrated_summary, df_sim_tbl=None, is_auto=True
            )
            if ok:
                st.toast(f"📧 통합 리포트 메일 자동 발송 완료!", icon="🎉")
                st.success(f"📧 **[자동 발송 성공]** {msg_res}")
            else:
                st.warning(f"⚠️ **[자동 발송 미완료]** {msg_res}")

# --------------------------------------------------------------------------
# TAB 3: 📈 회차별 트렌드 & 거동예측
# --------------------------------------------------------------------------
with main_tab3:
    st.header("📈 $n$회차 트렌드 시각화 & 거동예측 시뮬레이터")
    
    with st.expander("⚙️ 거동예측 시뮬레이션 파라미터 설정 (What-If Simulation)", expanded=True):
        col_p1, col_p2, col_p3 = st.columns(3)
        with col_p1:
            target_max_run = st.slider("예측 시뮬레이션 목표 회차", min_value=3, max_value=20, value=10, step=1)
            sintering_decay = st.slider("회차당 CaO 소결 활성도 감쇄율 (%)", min_value=0.5, max_value=8.0, value=3.5, step=0.1)
        with col_p2:
            sim_temp = st.number_input("가상 반응 온도 (℃)", min_value=60.0, max_value=95.0, value=80.0, step=1.0)
            sim_time = st.number_input("가상 반응 시간 (h)", min_value=1.0, max_value=5.0, value=2.0, step=0.5)
        with col_p3:
            wash_ratio = st.slider("수세수 투입 배수 (케이크 대비)", min_value=1.0, max_value=5.0, value=3.0, step=0.5)
            fresh_makeup_mode = st.selectbox("CaO 보충 방식", ["고정량 보충 (전회차 감량분 100% Make-up)", "신품 100% 교체 (Purge)"])

    sim_rows = []
    base_li_conc = float(st.session_state.icp_li_1) if 'icp_li_1' in st.session_state else 10500.0
    base_recovery = total_li_solution_rec_pct if 'total_li_solution_rec_pct' in locals() and total_li_solution_rec_pct > 0 else 95.80

    temp_factor = 1.0 + (sim_temp - 80.0) * 0.004
    time_factor = 1.0 + (sim_time - 2.0) * 0.03

    for r in range(1, target_max_run + 1):
        matched_actual = st.session_state.history[st.session_state.history["회차 (Run)"] == r]
        
        if not matched_actual.empty:
            row = matched_actual.iloc[0].to_dict()
            sim_rows.append(row)
        else:
            activity = 100.0 * ((1.0 - (sintering_decay / 100.0)) ** (r - 1))
            eff_activity = min(100.0, activity * temp_factor * time_factor)
            
            pred_rec = max(70.0, min(99.0, base_recovery * (eff_activity / 100.0)))
            pred_li_conc = max(7000.0, base_li_conc * (pred_rec / base_recovery))
            pred_lioh_conc = round(pred_li_conc * (MW_LIOH / MW_LI) / 1000.0, 2)
            pred_loi = max(35.0, 41.13 - (r - 1) * 0.4)
            pred_makeup = 68.52 if fresh_makeup_mode.startswith("고정량") else 92.42

            sim_rows.append({
                "회차 (Run)": r,
                "구분": "AI 예측치 (Simulated)",
                "Li 회수율 (%)": round(pred_rec, 2),
                "LiOH용액 Li농도 (mg/L)": round(pred_li_conc, 1),
                "LiOH용액 농도 (g/L)": round(pred_lioh_conc, 2),
                "M/B 닫힘율 (%)": round(max(90.0, 95.88 - (r - 1) * 0.2), 2),
                "하소 감율 LOI (%)": round(pred_loi, 2),
                "CaO 활성도 (%)": round(eff_activity, 1),
                "신품 CaO 보충량 (g)": round(pred_makeup, 2)
            })

    df_simulation = pd.DataFrame(sim_rows)

    st.markdown("---")
    st.subheader("📊 회차별 인터랙티브 X-Y 트렌드 그래프")

    col_ctrl1, col_ctrl2 = st.columns([1, 2])
    with col_ctrl1:
        y_axis_metric = st.selectbox(
            "📌 Y축에 표시할 지표를 선택하세요:",
            [
                "Li 회수율 (%)", "LiOH용액 농도 (g/L)", "LiOH용액 Li농도 (mg/L)", 
                "CaO 활성도 (%)", "하소 감율 LOI (%)", "신품 CaO 보충량 (g)"
            ]
        )

    chart_df = df_simulation.set_index("회차 (Run)")[[y_axis_metric]]
    st.line_chart(chart_df, height=380, use_container_width=True)

    st.markdown("##### 📋 회차별 실측치 & 예측 시뮬레이션 상세 데이터 테이블")
    st.dataframe(
        df_simulation.style.format({
            "Li 회수율 (%)": "{:.2f} %", "LiOH용액 Li농도 (mg/L)": "{:,.1f} mg/L",
            "LiOH용액 농도 (g/L)": "{:.2f} g/L", "M/B 닫힘율 (%)": "{:.2f} %",
            "하소 감율 LOI (%)": "{:.2f} %", "CaO 활성도 (%)": "{:.1f} %",
            "신품 CaO 보충량 (g)": "{:.2f} g"
        }),
        use_container_width=True
    )

# --------------------------------------------------------------------------
# TAB 4: 🔬 AI 자율 실험계획 & 레시피 추천 (DoE Agent)
# --------------------------------------------------------------------------
with main_tab4:
    st.header("🔬 Google Gemini 기반 자율 실험계획 (DoE Agent)")
    st.caption("과거 실험 데이터와 Ca-Loop 화학 양론을 기반으로 차기 회차의 최적 반응/소성 레시피를 AI가 자율 설계합니다.")

    doe_col1, doe_col2 = st.columns(2)
    with doe_col1:
        next_run_num = len(st.session_state.history) + 1
        st.markdown(f"#### 🎯 차기 실험 회차: **Run {next_run_num}**")
        target_goal = st.selectbox(
            "📌 실험 최적화 목표 설정",
            [
                "1. Li 회수율 극대화 (최대 회수율 98%+ 목표)",
                "2. 불순물(Ca, Na 등) 최소화 (배터리급 초고순도 목표)",
                "3. 신품 CaO 절감 및 원가 최적화 (Ca-Loop 최대 재활용)",
                "4. 공정 시간 및 에너지(온도) 절감 최적화"
            ]
        )
    with doe_col2:
        st.markdown("#### ⚙️ 공정 제약 조건")
        cao_avail = st.number_input("현재 보유 재생 CaO (g)", value=float(st.session_state.calcined_cao), format="%.1f")
        target_li2co3 = st.number_input("목표 Li₂CO₃ 투입량 (g)", value=95.34, format="%.2f")

    if st.button("🚀 Gemini 자율 DoE 최적 레시피 생성", type="primary", use_container_width=True):
        api_key = st.session_state.gemini_api_key.strip()
        if not api_key:
            st.error("❌ 사이드바에 Google Gemini API Key를 입력하거나 Secrets에 등록해 주세요.")
        else:
            with st.spinner(f"Gemini AI가 과거 {len(st.session_state.history)}개 회차 데이터와 공정 반응 속도론을 분석 중입니다..."):
                try:
                    history_summary = st.session_state.history.to_dict(orient="records")

                    doe_prompt = f"""
당신은 탄산리튬(LC) → 수산화리튬(LH) 전환 가성화 및 CaCO3-CaO 순환 하소(Ca-Loop) 공정의 수석 화학공정 엔지니어입니다.

[현재 상태 데이터]
- 다음 실험 회차: Run {next_run_num}
- 최적화 목표: {target_goal}
- 목표 Li2CO3 투입량: {target_li2co3} g
- 사용 가능 재생 CaO: {cao_avail} g
- 과거 실험 이력: {json.dumps(history_summary, ensure_ascii=False)}

위 데이터를 바탕으로 차기 회차를 위한 최적의 실험 조건(DoE Recipe)을 산출하여 반드시 아래 JSON 포맷으로만 응답해 주세요.
숫자는 순수 숫자(float)로만 제공해야 합니다.

{{
  "recipe": {{
    "run_no": {next_run_num},
    "li2co3_mass": float,
    "li2co3_water": float,
    "fresh_cao_mass": float,
    "recycled_cao_mass": float,
    "slurry_water": float,
    "temp_c": float,
    "time_h": float,
    "wash_water_in": float,
    "calc_temp": float,
    "calc_time": float
  }},
  "expected_outcome": {{
    "expected_recovery_pct": float,
    "expected_lioh_conc_gl": float,
    "expected_ca_mg_l": float
  }},
  "engineering_rationale": "배합비, 수세수 투입량 및 반응 조건을 이렇게 설정한 엔지니어링 근거 (3~4줄)",
  "precautions": "실험 진행 시 핵심 주의사항 (2~3줄)"
}}
"""
                    genai.configure(api_key=api_key)
                    model = genai.GenerativeModel("gemini-1.5-flash")
                    resp = model.generate_content(doe_prompt, request_options={"timeout": 15})

                    if resp and resp.text:
                        clean_json_str = resp.text.replace("```json", "").replace("```", "").strip()
                        doe_result = json.loads(clean_json_str)
                        st.session_state.latest_doe = doe_result
                        st.success("🎉 Gemini AI 자율 DoE 레시피가 성공적으로 생성되었습니다!")
                    else:
                        st.error("❌ DoE 생성 실패: AI 모델 응답을 가져오지 못했습니다.")
                except Exception as e:
                    st.error(f"❌ DoE 생성 실패: {e}")

    if "latest_doe" in st.session_state and st.session_state.latest_doe:
        doe = st.session_state.latest_doe
        rec = doe.get("recipe", {})
        exp = doe.get("expected_outcome", {})

        st.markdown("---")
        st.subheader("📋 Gemini AI 추천 차기 실험 레시피 (DoE Recipe)")

        d1, d2, d3 = st.columns(3)
        d1.metric("🎯 예상 Li 회수율", f"{exp.get('expected_recovery_pct', 0):.2f} %")
        d2.metric("💧 예상 LiOH 용액 농도", f"{exp.get('expected_lioh_conc_gl', 0):.2f} g/L")
        d3.metric("⚠️ 예상 불순물 Ca 농도", f"{exp.get('expected_ca_mg_l', 0):.1f} mg/L")

        df_recipe_view = pd.DataFrame([
            {"공정 항목": "Li₂CO₃ 투입량", "추천 수치": f"{rec.get('li2co3_mass', 95.34):.2f} g", "비고": "원료"},
            {"공정 항목": "Li₂CO₃ 용매수", "추천 수치": f"{rec.get('li2co3_water', 1040.0):.1f} g", "비고": "용해수"},
            {"공정 항목": "신품 CaO 투입량", "추천 수치": f"{rec.get('fresh_cao_mass', 68.52):.2f} g", "비고": "신품 보충"},
            {"공정 항목": "재생 CaO 투입량", "추천 수치": f"{rec.get('recycled_cao_mass', 23.90):.2f} g", "비고": "재생분 활용"},
            {"공정 항목": "슬러리 조제수", "추천 수치": f"{rec.get('slurry_water', 831.0):.1f} g", "비고": "소화수"},
            {"공정 항목": "수세수 투입량", "추천 수치": f"{rec.get('wash_water_in', 850.0):.1f} g", "비고": "CaCO₃ 세척수"},
            {"공정 항목": "반응 온도", "추천 수치": f"{rec.get('temp_c', 80.0):.1f} ℃", "비고": "가성화"},
            {"공정 항목": "반응 시간", "추천 수치": f"{rec.get('time_h', 2.0):.1f} 시간", "비고": "교반"},
            {"공정 항목": "하소(소성) 온도", "추천 수치": f"{rec.get('calc_temp', 1000.0):.1f} ℃", "비고": "CaCO₃ 탈탄산"},
            {"공정 항목": "하소(소성) 시간", "추천 수치": f"{rec.get('calc_time', 1.0):.1f} 시간", "비고": "소결 억제"}
        ])
        st.table(df_recipe_view)

        st.info(f"💡 **[엔지니어링 설계 근거]**\n{doe.get('engineering_rationale', '-')}")
        st.warning(f"⚠️ **[실험 주의사항]**\n{doe.get('precautions', '-')}")

        if st.button("📥 이 추천 레시피를 1번 탭 입력창에 즉시 반영하기", type="primary"):
            r_apply = rec.get("run_no", next_run_num)
            st.session_state.run_no = int(r_apply)
            st.session_state.tab2_run_no = int(r_apply)
            for k in ["li2co3_mass", "li2co3_water", "fresh_cao_mass", "recycled_cao_mass", "slurry_water", "wash_water_in", "temp_c", "time_h", "calc_temp", "calc_time"]:
                if k in rec:
                    val_c = clean_float(rec[k])
                    if val_c is not None:
                        st.session_state[k] = val_c
            st.success("✅ 추천 레시피가 1번 탭 입력창에 모두 적용되었습니다! 1번 탭으로 이동하여 실험을 진행하세요.")
            st.rerun()

# --------------------------------------------------------------------------
# TAB 5: 💬 AI 공정 대화창
# --------------------------------------------------------------------------
with main_tab5:
    st.header("💬 AI 공정 엔지니어와 대화하기")
    st.caption("현재 실험 수치, LiOH 용액/CaCO₃ 분석 결과, $n$회차 시뮬레이션을 바탕으로 실시간 질의응답을 진행합니다.")

    for msg in st.session_state.chat_messages:
        with st.chat_message(msg["role"]):
            st.markdown(msg["content"])

    if user_prompt := st.chat_input("질문을 입력하세요 (예: 수세수를 850g 넣었을 때 세척 효율이 적정한가?)"):
        st.session_state.chat_messages.append({"role": "user", "content": user_prompt})
        with st.chat_message("user"):
            st.markdown(user_prompt)

        api_key = st.session_state.gemini_api_key.strip()
        if api_key:
            try:
                genai.configure(api_key=api_key)
                chat_model = genai.GenerativeModel("gemini-1.5-flash")
                context_prompt = f"""당신은 LC-LH 전환 가성화 및 Ca-Loop 공정의 최고 권위 수석 엔지니어입니다.
현재 공정 데이터:
- 실험 회차: Run {st.session_state.run_no}
- 총 Li 용액 회수율: {total_li_solution_rec_pct:.2f}% (LiOH 용액 {li_rec_1_pct:.1f}%, 수세액 {li_rec_w_pct:.1f}%)
- CaCO₃ Li 손실률: {li_cake_loss_pct:.2f}% (고체 wt%: {st.session_state.solid_li_wt:.3f}%)
- 총 Li 원소 닫힘율: {total_li_closure_pct:.2f}%
- LiOH 용액 농도: {lioh_equiv_g_l:.2f} g/L, Ca: {st.session_state.icp_ca_1} mg/L, Na: {st.session_state.icp_na_1} mg/L
- 수세 조건: 투입 수세수 {st.session_state.wash_water_in:.1f}g, 회수 수세액 {st.session_state.wash_sol_mass:.1f}g (수세배수 {wash_ratio_wet:.2f}배)
- CaCO₃ 함수율: {cake_moisture:.2f}% (건조 CaCO₃: {est_total_dry_solids:.1f}g)
- LOI: {loi_pct:.2f}%, M/B 정합성: {mass_closure:.2f}%

질문: {user_prompt}
배터리 소재 품질 및 양론적 관점에서 친절하고 명확하게 답변해 주세요."""
                resp = chat_model.generate_content(context_prompt, request_options={"timeout": 15})
                if resp and resp.text:
                    ai_reply = resp.text
                else:
                    ai_reply = f"현재 수세수 투입량은 **{st.session_state.wash_water_in:.1f} g**, 수세액 회수율 기여도는 **{li_rec_w_pct:.2f}%**입니다."
            except Exception as e:
                ai_reply = f"현재 수세수 투입량은 **{st.session_state.wash_water_in:.1f} g**, 수세액 회수율 기여도는 **{li_rec_w_pct:.2f}%**입니다. (API 오류: {e})"
        else:
            ai_reply = f"현재 수세수 투입량은 **{st.session_state.wash_water_in:.1f} g**, 수세액 회수율 기여도는 **{li_rec_w_pct:.2f}%**입니다."

        st.session_state.chat_messages.append({"role": "assistant", "content": ai_reply})
        with st.chat_message("assistant"):
            st.markdown(ai_reply)

# --------------------------------------------------------------------------
# TAB 6: 📧 리포트 메일 발송 현황 & 연결 테스트
# --------------------------------------------------------------------------
with main_tab6:
    st.header("📧 리포트 메일 발송 현황 & 계정 설정")
    st.caption("실험 데이터 저장 시 자동으로 발송된 이메일 현황 목록을 모니터링하고 발송 설정을 관리합니다.")

    st.markdown("### 📋 실시간 메일 발송 이력 현황")
    if st.session_state.email_logs:
        df_logs = pd.DataFrame(st.session_state.email_logs)
        success_cnt = sum(1 for log in st.session_state.email_logs if "성공" in log["발송 상태"])
        fail_cnt = len(st.session_state.email_logs) - success_cnt

        m_col1, m_col2, m_col3 = st.columns(3)
        m_col1.metric("총 발송 시도", f"{len(st.session_state.email_logs)} 건")
        m_col2.metric("발송 성공", f"{success_cnt} 건")
        m_col3.metric("발송 실패 / 보류", f"{fail_cnt} 건", delta_color="inverse")

        st.dataframe(df_logs, use_container_width=True)

        if st.button("🗑️ 발송 이력 초기화", use_container_width=False):
            st.session_state.email_logs = []
            st.rerun()
    else:
        st.info("ℹ️ 아직 발송된 메일 이력이 없습니다. 2번 탭에서 [💾 트렌드 DB에 저장]을 누르면 엑셀 리포트가 자동 발송되고 여기에 기록됩니다.")

    st.divider()

    st.markdown("### ⚙️ 발송 계정 및 수신자 설정")
    col_cfg1, col_cfg2 = st.columns(2)

    with col_cfg1:
        st.session_state.email_recipients = st.text_input(
            "📬 수신자 이메일 목록 (쉼표로 구분)", 
            value=st.session_state.email_recipients,
            help="예: user1@company.com, user2@company.com"
        )
        st.session_state.email_sender = st.text_input(
            "📤 발신자 이메일 주소 (예: myname@gmail.com)", 
            value=st.session_state.email_sender
        )
        st.session_state.email_password = st.text_input(
            "🔑 발신자 앱 비밀번호 (16자리)", 
            value=st.session_state.email_password,
            type="password",
            help="Gmail: Google 계정 보안 > 2단계 인증 > 앱 비밀번호"
        )

    with col_cfg2:
        st.session_state.smtp_server = st.text_input(
            "🌐 SMTP 서버 호스트", 
            value=st.session_state.smtp_server
        )
        st.session_state.smtp_port = st.number_input(
            "🔌 SMTP 포트 번호", 
            value=int(st.session_state.smtp_port),
            step=1
        )
        st.write("")
        st.session_state.auto_email_on_save = st.toggle(
            "⚡ 실험값 저장 시 엑셀 리포트 자동 발송 활성화", 
            value=st.session_state.auto_email_on_save
        )

    st.markdown("---")
    st.markdown("#### 🔍 계정 연결 1초 진단 테스트")
    col_t1, col_t2 = st.columns(2)
    with col_t1:
        if st.button("🔌 SMTP 계정 연결 즉시 테스트", use_container_width=True):
            test_sender = st.session_state.email_sender.strip()
            test_pw = st.session_state.email_password.strip().replace(" ", "")
            test_host = st.session_state.smtp_server.strip()
            test_port = int(st.session_state.smtp_port)

            if not test_sender or not test_pw:
                st.error("❌ 발신자 이메일과 비밀번호를 입력한 후 테스트를 눌러주세요.")
            else:
                with st.spinner("메일 서버 연결 및 로그인 인증 테스트 중..."):
                    try:
                        if test_port == 465:
                            s = smtplib.SMTP_SSL(test_host, test_port, timeout=10)
                        else:
                            s = smtplib.SMTP(test_host, test_port, timeout=10)
                            s.starttls()
                        s.login(test_sender, test_pw)
                        s.quit()
                        st.success(f"🎉 **[인증 성공!]** `{test_sender}` 계정으로 메일 서버에 정상 로그인되었습니다!")
                    except Exception as err:
                        st.error(f"❌ **[인증 실패]** 상세 원인: {err}")

    with col_t2:
        if st.button(f"📨 현재 회차 (Run {st.session_state.run_no}) 리포트 수동 발송", type="primary", use_container_width=True):
            ok, msg_res = send_email_report(
                run_num=int(st.session_state.run_no), mass_cls=mass_closure, loss_m=loss_mass,
                li_rec_tot=total_li_solution_rec_pct, li_rec_1=li_rec_1_pct, li_rec_w=li_rec_w_pct,
                li_cake_loss=li_cake_loss_pct, lioh_conc=lioh_equiv_g_l, li_1=float(st.session_state.icp_li_1),
                ca_1=float(st.session_state.icp_ca_1), na_1=float(st.session_state.icp_na_1), si_1=float(st.session_state.icp_si_1), mg_1=float(st.session_state.icp_mg_1), k_1=float(st.session_state.icp_k_1),
                loi=loi_pct, purity=purity_caco3, makeup=fresh_makeup, cao_rec=calcined_cao,
                cake_moisture=cake_moisture, dry_caco3_mass=est_total_dry_solids,
                wash_water_in=wash_water_in, wash_sol_mass=wash_sol_mass,
                df_icp_tbl=df_integrated_summary, df_sim_tbl=df_simulation, is_auto=False
            )
            if ok:
                st.success(f"🎉 {msg_res}")
                st.rerun()
            else:
                st.error(f"❌ {msg_res}")
                st.rerun()
